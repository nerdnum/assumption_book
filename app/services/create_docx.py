import json
import os
from io import BytesIO
from datetime import datetime

from lxml import etree
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.workbook.views import BookView
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from app.config import get_config
from app.html2docx.htmldocx import HtmlToDocx
from app.services.database import sessionmanager
from app.services.utils import pretty_print
from app.sqlalchemy_models.components_sql import Component as SqlComponent
from app.sqlalchemy_models.user_project_role_sql import Project as SqlProject


def processHtml(html):
    tree = etree.HTML(html)
    images = tree.xpath("//img")
    cwd = os.getcwd()
    if len(images) > 0:
        for image in images:
            src = image.attrib["src"]
            _, path = src.split("/static/document_images/")
            store_path = os.path.join(cwd, "app", "static", "document_images", path)
            image.attrib["src"] = store_path
    html_string = (
        etree.tostring(tree, encoding="utf-8").decode("utf-8").replace("\n", "")
    )
    return html_string


async def create_project_docx(spec):

    async with sessionmanager.session() as db:
        project = await SqlProject.get_project_by_id(db, spec.project_id)
        html = f'<p style="docx-style: Title"><strong>{project.title}</strong></p>'
        for component_spec in spec.components:
            component_html = await SqlComponent.get_html_by_id(db, component_spec.id)
            html += component_html

        final_html = processHtml(html)
        parser = HtmlToDocx()
        parser.table_style = "Grid Table 6 Colorful Accent 1"
        cwd = os.getcwd()
        template_path = os.path.join(cwd, "app", "static", "template.docx")
        docx = parser.parse_html_string(final_html, template=template_path)
        suffix = datetime.now().strftime("%M%S")
        doc_name = f'{project.title.replace(" ", "_").lower()}_{suffix}'
        store_path = os.path.join(cwd, "app", "static", "docx", f"{doc_name}.docx")
        docx.save(store_path)
        store_path = os.path.join(cwd, "app", "static", "docx", f"{doc_name}.html")
        url_path = os.path.join("static", "docx", f"{doc_name}.docx")
        with open(store_path, "w", encoding="utf-8") as file:
            file.write(final_html)

        return json.dumps(
            {"status": "success", "name": f"{doc_name}.docx", "url": url_path}
        )


def try_to_make_number(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return value


def get_row_data(parameters_json):
    rows = []
    rows.append(["TITLE", "VALUE", "UNIT", "SOURCE", "COMMENT"])
    for data in parameters_json:
        row = [
            data.get("title"),
            try_to_make_number(data.get("value")),
            data.get("unit").get("value").get("displayUnit"),
            data.get("source"),
            data.get("comment"),
        ]
        rows.append(row)
    return rows


# BCCBE2
table_fill_darker = PatternFill(
    start_color="BCCBE2", end_color="BCCBE2", fill_type="solid"
)
table_fill_lighter = PatternFill(
    start_color="DEE6F0", end_color="DEE6F0", fill_type="solid"
)
wrap_text = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(horizontal="center")
table_cell_side = Side(style="thin", color="9BB2D4")
thick = Side(style="thick", color="9BB2D4")
normal_borders = Border(
    left=table_cell_side,
    right=table_cell_side,
    top=table_cell_side,
    bottom=table_cell_side,
)
bottom_thick_borders = Border(
    left=table_cell_side,
    right=table_cell_side,
    top=table_cell_side,
    bottom=thick,
)

heading = Font(bold=True, name="Arial", size=12)


async def create_project_xlsx(spec):
    async with sessionmanager.session() as db:
        project = await SqlProject.get_project_by_id(db, spec.project_id)
        workbook = Workbook()
        view = [BookView(xWindow=0, yWindow=0, windowWidth=27210, windowHeight=23310)]
        workbook.views = view
        del workbook["Sheet"]  # Remove the default sheet

        spec_count = 0

        for component_spec in spec.components:
            component = await SqlComponent.get_by_id(db, component_spec.id)

            component_json = await SqlComponent.get_parameters_json_by_component_id(
                db, component_spec.id
            )
            if len(component_json["documents"]) > 0:
                spec_count += 1
                workbook.create_sheet(title=component_json["title"])
                worksheet = workbook[component_json["title"]]
                worksheet.column_dimensions["A"].width = 30
                worksheet.column_dimensions["B"].width = 12
                worksheet.column_dimensions["C"].width = 12
                worksheet.column_dimensions["D"].width = 30
                worksheet.column_dimensions["E"].width = 30
                row_count = 0
                for block in component_json["documents"]:
                    component_id = block.get("component_id", "")
                    if block.get("type") == "parameters":
                        title = (
                            f"Parameters Scope: {block['document'].get('scope', '')}"
                        )
                    else:  # is interface
                        interfaceDef = block["document"].get("interfacedComponent", {})
                        if interfaceDef:
                            if component_id == interfaceDef.get("componentOneId"):
                                title = f"Interface: {interfaceDef.get('componentOneTitle', '')} with {interfaceDef.get('componentTwoTitle', '')}"
                            else:
                                title = f"Interface: {interfaceDef.get('componentTwoTitle', '')} with {interfaceDef.get('componentOneTitle', '')}"

                    description = block["document"].get("description", "")
                    worksheet.append(
                        [
                            title,
                        ]
                    )

                    row_count += 1
                    worksheet.cell(row=row_count, column=1).font = heading
                    worksheet.append(
                        [
                            "Description",
                        ]
                    )
                    row_count += 1
                    worksheet.cell(row=row_count, column=1).font = heading
                    worksheet.append(
                        [
                            description,
                        ]
                    )
                    row_count += 1
                    worksheet.merge_cells(
                        start_row=row_count,
                        start_column=1,
                        end_row=row_count,
                        end_column=5,
                    )
                    worksheet.row_dimensions[row_count].height = 60
                    worksheet.cell(row=row_count, column=1).alignment = wrap_text

                    data_rows = get_row_data(block["document"].get("parameters"))

                    is_first_row = True
                    table_row_count = 0
                    for row in data_rows:
                        worksheet.append(row)
                        row_count += 1
                        for col in range(1, 6):
                            cell = worksheet.cell(row=row_count, column=col)
                            cell.alignment = wrap_text
                            if is_first_row:
                                cell.font = Font(bold=True)
                                cell.border = bottom_thick_borders
                            else:
                                cell.border = normal_borders
                            cell.fill = (
                                table_fill_darker
                                if table_row_count % 2 == 0
                                else table_fill_lighter
                            )
                            if col in [2, 3]:
                                cell.alignment = center
                        is_first_row = False
                        table_row_count += 1
                    worksheet.append([])
                    row_count += 1
        # Set the first sheet as active
        if spec_count > 0:
            workbook.active = 0
        else:
            worksheet = workbook.create_sheet(title="No Components")
            worksheet.append(
                [
                    "No components with parameter or interface tables found for this project."
                ]
            )
            worksheet.append(
                [
                    "Please add parameter or interface tables to components to see content in this spreadsheet document."
                ]
            )
            worksheet.append([" "])
        cwd = os.getcwd()
        suffix = datetime.now().strftime("%M%S")
        doc_name = f'{project.title.replace(" ", "_").lower()}_{suffix}'
        store_path = os.path.join(cwd, "app", "static", "xlsx", f"{doc_name}.xlsx")
        workbook.save(store_path)
        url_path = os.path.join("static", "xlsx", f"{doc_name}.xlsx")

    return json.dumps(
        {"status": "success", "name": f"{doc_name}.xlsx", "url": url_path}
    )
